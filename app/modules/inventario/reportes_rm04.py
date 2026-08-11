"""Registro Mensual SUNAT (RM04): consulta, armado del libro y descarga.

El libro tiene cuatro hojas y dos propósitos distintos:

* `RDO` reproduce la plantilla oficial que el laboratorio presenta a SUNAT:
  una fila por frasco, cabecera de cinco filas combinadas, stock al primer
  y al último día del periodo, y el movimiento del mes con su tipo de
  registro y su curso. Es la hoja que se declara.

* `CONTROL`, `MOVIMIENTOS` y `RECONCILIACION` son el contrato interno de
  EN-004 y sirven de sustento: dejan a la vista el folio, la trazabilidad
  y el cuadre que la plantilla oficial resume en dos columnas.

Todo se calcula en gramos y solo se convierte a kilogramos al escribir la
celda, una sola vez, como manda EN-004.
"""

import datetime as dt
import hashlib
import io
import re
from decimal import Decimal
from typing import Annotated, Any

import openpyxl
from fastapi import APIRouter, Depends, Query, Response
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from psycopg import Connection

from app.dependencies import get_connection, get_current_user
from app.modules.auth.schemas import CurrentUser

router = APIRouter(prefix="/api/inventario/reportes", tags=["Inventario y consumo"])

ESTABLECIMIENTO = "ULIMA-DOCIMASIA"

# La plantilla oficial codifica el movimiento con un catálogo cerrado. Un
# tipo que no esté en él NO se fuerza a 1 ni a 4: se escribe tal cual para
# que salte a la vista en la revisión, en vez de declararse mal.
TIPO_REGISTRO_SUNAT = {"ENTRADA": "1", "SALIDA": "4"}


def _grado(texto: str | None) -> str:
    """Reduce el grado del lote al vocabulario de la plantilla (PA / ACS)."""
    if not texto:
        return ""
    normalizado = texto.upper()
    if "ACS" in normalizado:
        return "ACS"
    # "pro analysi", "for analysis", "zur Analyse", "reactivo analítico":
    # todas son la misma calidad, escrita en cuatro idiomas distintos.
    if any(marca in normalizado
           for marca in ("P.A", "Q.P", "ANALYSI", "ANALYSE", "ANALITIC", "ANALÍTIC")):
        return "PA"
    return texto[:12]


def _unidad_comercial(tipo_envase: str | None) -> str:
    """Traduce el envase del maestro al código comercial de la plantilla."""
    if not tipo_envase:
        return ""
    normalizado = tipo_envase.upper()
    if "BOTELLA" in normalizado:
        return "BOT"
    if "SACO" in normalizado or "BOLSA" in normalizado:
        return "SACO"
    return "ENV"


def _kg(gramos: Any) -> float:
    """Convierte a kilogramos una sola vez, al escribir la celda."""
    if gramos is None:
        return 0.0
    return float(Decimal(str(gramos)) / Decimal("1000"))


def _presentaciones(gramos: Any, equivalencia_g: Any) -> float | str:
    """Expresa un saldo en fracción de envase; vacío si no hay equivalencia."""
    if gramos is None or not equivalencia_g or Decimal(str(equivalencia_g)) == 0:
        return ""
    return float(Decimal(str(gramos)) / Decimal(str(equivalencia_g)))


# ── Consultas ──────────────────────────────────────────────────────────────

# Una fila por frasco y por movimiento del periodo. El frasco sin movimiento
# aparece igual, con las columnas de movimiento vacías: la plantilla declara
# existencias, no solo operaciones.
SQL_RDO = """
WITH periodo AS (
    SELECT to_date(%(periodo)s, 'YYYY-MM') AS inicio,
           (to_date(%(periodo)s, 'YYYY-MM') + INTERVAL '1 month'
             - INTERVAL '1 day')::date AS fin
),
saldo_inicial AS (
    SELECT DISTINCT ON (k.id_frasco) k.id_frasco, k.saldo_resultante_g
      FROM iqbf.kardex k, periodo p
     WHERE k.fecha_operacion < p.inicio
     ORDER BY k.id_frasco, k.fecha_operacion DESC, k.id_movimiento DESC
),
saldo_final AS (
    SELECT DISTINCT ON (k.id_frasco) k.id_frasco, k.saldo_resultante_g
      FROM iqbf.kardex k, periodo p
     WHERE k.fecha_operacion <= p.fin
     ORDER BY k.id_frasco, k.fecha_operacion DESC, k.id_movimiento DESC
)
SELECT i.nombre_comercial,
       l.grado_pureza,
       pr.codigo_bf_sunat,
       f.id_frasco,
       pr.tipo_envase,
       pr.capacidad,
       pr.unidad,
       pr.equivalencia_g,
       COALESCE(si.saldo_resultante_g, 0) AS saldo_inicial_g,
       COALESCE(sf.saldo_resultante_g, 0) AS saldo_final_g,
       k.fecha_operacion,
       k.tipo_movimiento,
       k.curso,
       k.cantidad_g
  FROM iqbf.frasco f
  JOIN iqbf.lote l         ON l.id_lote = f.id_lote
  JOIN iqbf.presentacion pr ON pr.id_presentacion = l.id_presentacion
  JOIN iqbf.insumo i       ON i.id_insumo = pr.id_insumo
  CROSS JOIN periodo p
  LEFT JOIN saldo_inicial si ON si.id_frasco = f.id_frasco
  LEFT JOIN saldo_final  sf ON sf.id_frasco = f.id_frasco
  LEFT JOIN iqbf.kardex   k  ON k.id_frasco = f.id_frasco
       AND k.fecha_operacion BETWEEN p.inicio AND p.fin
 WHERE f.estado <> 'DADO_DE_BAJA'
 ORDER BY i.nombre_comercial, f.id_frasco, k.fecha_operacion, k.id_movimiento
"""

SQL_MOVIMIENTOS = """
SELECT k.id_movimiento::text            AS folio,
       k.fecha_operacion,
       k.tipo_movimiento,
       k.motivo,
       pr.id_insumo                     AS insumo_codigo,
       i.nombre_comercial,
       pr.codigo_bf_sunat,
       pr.id_presentacion,
       i.tipo                           AS estado_fisico,
       l.numero_lote,
       k.id_frasco,
       k.unidad_registrada,
       k.cantidad_registrada,
       k.densidad_aplicada,
       k.cantidad_g,
       k.saldo_resultante_g,
       k.curso,
       COALESCE(inv.nombre, '')         AS responsable,
       COALESCE(u.codigo_institucional, u.nombre, '') AS registrado_por
  FROM iqbf.kardex k
  JOIN iqbf.frasco f        ON f.id_frasco = k.id_frasco
  JOIN iqbf.lote l          ON l.id_lote = f.id_lote
  JOIN iqbf.presentacion pr ON pr.id_presentacion = l.id_presentacion
  JOIN iqbf.insumo i        ON i.id_insumo = pr.id_insumo
  LEFT JOIN iqbf.investigador inv ON inv.id_investigador = f.id_investigador
  LEFT JOIN iqbf.usuario u        ON u.id_usuario = k.registrado_por
 WHERE to_char(k.fecha_operacion, 'YYYY-MM') = %(periodo)s
 ORDER BY k.fecha_operacion, k.id_movimiento
"""

SQL_RECONCILIACION = """
WITH periodo AS (
    SELECT to_date(%(periodo)s, 'YYYY-MM') AS inicio,
           (to_date(%(periodo)s, 'YYYY-MM') + INTERVAL '1 month'
             - INTERVAL '1 day')::date AS fin
),
saldo_inicial AS (
    SELECT DISTINCT ON (k.id_frasco) k.id_frasco, k.saldo_resultante_g
      FROM iqbf.kardex k, periodo p
     WHERE k.fecha_operacion < p.inicio
     ORDER BY k.id_frasco, k.fecha_operacion DESC, k.id_movimiento DESC
),
base AS (
    SELECT pr.codigo_bf_sunat,
           i.nombre_comercial,
           f.id_frasco,
           COALESCE(si.saldo_resultante_g, 0) AS inicial_g,
           f.peso_neto_actual_g               AS fisico_g
      FROM iqbf.frasco f
      JOIN iqbf.lote l          ON l.id_lote = f.id_lote
      JOIN iqbf.presentacion pr ON pr.id_presentacion = l.id_presentacion
      JOIN iqbf.insumo i        ON i.id_insumo = pr.id_insumo
      LEFT JOIN saldo_inicial si ON si.id_frasco = f.id_frasco
     WHERE f.estado <> 'DADO_DE_BAJA'
),
movs AS (
    SELECT f.id_frasco,
           SUM(k.cantidad_g) FILTER (WHERE k.tipo_movimiento = 'ENTRADA') AS entradas_g,
           SUM(k.cantidad_g) FILTER (WHERE k.tipo_movimiento = 'SALIDA')  AS salidas_g,
           SUM(k.cantidad_g) FILTER (WHERE k.tipo_movimiento
                                          NOT IN ('ENTRADA', 'SALIDA'))   AS ajustes_g
      FROM iqbf.kardex k
      JOIN iqbf.frasco f ON f.id_frasco = k.id_frasco
      CROSS JOIN periodo p
     WHERE k.fecha_operacion BETWEEN p.inicio AND p.fin
     GROUP BY f.id_frasco
)
SELECT b.codigo_bf_sunat,
       max(b.nombre_comercial)              AS nombre_comercial,
       count(*)                             AS frascos,
       SUM(b.inicial_g)                     AS inicial_g,
       COALESCE(SUM(m.entradas_g), 0)       AS entradas_g,
       COALESCE(SUM(m.salidas_g), 0)        AS salidas_g,
       COALESCE(SUM(m.ajustes_g), 0)        AS ajustes_g,
       COALESCE(SUM(b.fisico_g), 0)         AS fisico_g
  FROM base b
  LEFT JOIN movs m ON m.id_frasco = b.id_frasco
 GROUP BY b.codigo_bf_sunat
 ORDER BY b.codigo_bf_sunat NULLS LAST
"""


# ── Armado del libro ───────────────────────────────────────────────────────

_FUENTE_TITULO = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_RELLENO_TITULO = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
_BORDE = Border(*(Side(style="thin", color="B4C6E7"),) * 4)
_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _cabecera_rdo(ws, periodo: str) -> None:
    """Reproduce la cabecera de cinco filas combinadas de la plantilla oficial."""
    inicio = dt.date.fromisoformat(f"{periodo}-01")
    siguiente = (inicio.replace(day=28) + dt.timedelta(days=4)).replace(day=1)
    fin = siguiente - dt.timedelta(days=1)
    meses = ("ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO",
             "AGOSTO", "SETIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE")
    mes = meses[inicio.month - 1]

    # Columnas que ocupan las cinco filas de alto.
    simples = {
        "A": "N°",
        "B": "Nombre Comercial del Producto",
        "C": "Grado del Producto",
        "D": "Alta SUNAT",
        "E": "Código Interno Laboratorio",
        "F": "Unidad Comercial",
        "G": "Cantidad de unidad física",
        "H": "Unidad de medida de control",
        "I": "Cantidad neto / unidad (peso neto de botella)",
    }
    for col, titulo in simples.items():
        ws.merge_cells(f"{col}1:{col}5")
        ws[f"{col}1"] = titulo

    ws.merge_cells("J1:K1")
    ws["J1"] = "Stock existente al"
    ws.merge_cells("J2:K2")
    ws["J2"] = f"01 DE {mes} DEL {inicio.year}"
    ws.merge_cells("J3:J5")
    ws["J3"] = "Nº Botellas/ sacos"
    ws.merge_cells("K3:K5")
    ws["K3"] = "Cant. KG"

    ws.merge_cells("L1:P1")
    ws["L1"] = "Movimiento del reactivo"
    ws.merge_cells("L2:P2")
    ws["L2"] = mes
    ws.merge_cells("L3:L5")
    ws["L3"] = "Fecha"
    ws.merge_cells("M3:M4")
    ws["M3"] = "Tipo de Registro"
    ws["M5"] = "1: INGRESO    4: USO"
    ws.merge_cells("N3:N5")
    ws["N3"] = "Curso"
    ws.merge_cells("O3:P3")
    ws["O3"] = "Consumo"
    ws.merge_cells("O4:O5")
    ws["O4"] = "Cantidad Neta (Kg)"
    ws.merge_cells("P4:P5")
    ws["P4"] = "Cantidad de Presentaciones (sacos/ botellas)"

    ws.merge_cells("Q1:R1")
    ws["Q1"] = "Stock existente al"
    ws.merge_cells("Q2:R2")
    ws["Q2"] = f"{fin.day} DE {mes} DEL {fin.year}"
    ws.merge_cells("Q3:Q5")
    ws["Q3"] = "CANT. Kg"
    ws.merge_cells("R3:R5")
    ws["R3"] = "Nº Botellas/ Sacos"
    ws.merge_cells("S1:S5")
    ws["S1"] = "I * J = K"

    for fila in ws.iter_rows(min_row=1, max_row=5, max_col=19):
        for celda in fila:
            celda.font = _FUENTE_TITULO
            celda.fill = _RELLENO_TITULO
            celda.alignment = _CENTRO
            celda.border = _BORDE

    anchos = {"A": 5, "B": 30, "C": 12, "D": 12, "E": 22, "F": 10, "G": 11,
              "H": 12, "I": 14, "J": 14, "K": 12, "L": 12, "M": 14, "N": 18,
              "O": 15, "P": 16, "Q": 12, "R": 14, "S": 14}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho
    ws.freeze_panes = "A6"


def _hoja_rdo(wb, periodo: str, filas: list[dict[str, Any]]):
    ws = wb.create_sheet("RDO", 0)
    _cabecera_rdo(ws, periodo)

    numero = 0
    frasco_previo = None
    for fila in filas:
        # La numeración sigue al frasco, no a la fila: un frasco con dos
        # movimientos en el mes ocupa dos renglones bajo el mismo N°.
        if fila["id_frasco"] != frasco_previo:
            numero += 1
            frasco_previo = fila["id_frasco"]

        equivalencia = fila["equivalencia_g"]
        tipo = fila["tipo_movimiento"]
        es_consumo = tipo == "SALIDA"

        ws.append([
            numero,
            fila["nombre_comercial"],
            _grado(fila["grado_pureza"]),
            fila["codigo_bf_sunat"] or "",
            fila["id_frasco"],
            _unidad_comercial(fila["tipo_envase"]),
            float(fila["capacidad"]) if fila["capacidad"] is not None else "",
            fila["unidad"] or "",
            _kg(equivalencia),
            _presentaciones(fila["saldo_inicial_g"], equivalencia),
            _kg(fila["saldo_inicial_g"]),
            fila["fecha_operacion"] or "",
            TIPO_REGISTRO_SUNAT.get(tipo, tipo or ""),
            fila["curso"] or "",
            _kg(fila["cantidad_g"]) if es_consumo else "",
            _presentaciones(fila["cantidad_g"], equivalencia) if es_consumo else 0,
            _kg(fila["saldo_final_g"]),
            _presentaciones(fila["saldo_final_g"], equivalencia),
            _kg(equivalencia) * float(_presentaciones(fila["saldo_inicial_g"], equivalencia) or 0),
        ])

    for fila in ws.iter_rows(min_row=6, max_col=19):
        for celda in fila:
            celda.border = _BORDE
    # Los códigos conservan sus ceros a la izquierda: se escriben como texto.
    for renglon in range(6, ws.max_row + 1):
        ws.cell(renglon, 4).number_format = "@"
        ws.cell(renglon, 12).number_format = "DD/MM/YYYY"
    return ws


def _hoja_control(wb, periodo: str, usuario: str, huella: str, filas: int) -> None:
    ws = wb.create_sheet("CONTROL")
    ws.append(["Campo", "Valor"])
    generado = dt.datetime.now(dt.timezone.utc).astimezone().isoformat(timespec="seconds")
    for campo, valor in (
        ("periodo", periodo),
        ("establecimiento_codigo", ESTABLECIMIENTO),
        ("tipo_reporte", "RM04_SUNAT"),
        ("version", "1"),
        ("generado_en", generado),
        ("generado_por", usuario),
        ("filas_declaradas", str(filas)),
        ("hash_contenido", huella),
        ("estado", "BORRADOR"),
    ):
        ws.append([campo, valor])
    for celda in ws[1]:
        celda.font = _FUENTE_TITULO
        celda.fill = _RELLENO_TITULO
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 72


def _hoja_tabla(wb, titulo: str, cabeceras: list[str], filas: list[list[Any]]) -> None:
    ws = wb.create_sheet(titulo)
    ws.append(cabeceras)
    for fila in filas:
        ws.append(fila)
    for celda in ws[1]:
        celda.font = _FUENTE_TITULO
        celda.fill = _RELLENO_TITULO
        celda.alignment = _CENTRO
    for indice in range(1, len(cabeceras) + 1):
        ws.column_dimensions[get_column_letter(indice)].width = 20
    ws.freeze_panes = "A2"


def generar_excel_rm04(
    periodo: str,
    rdo: list[dict[str, Any]],
    movimientos: list[dict[str, Any]],
    reconciliacion: list[dict[str, Any]],
    usuario: str = "SISTEMA",
) -> bytes:
    """Arma el libro RM04 completo y lo devuelve en bytes."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _hoja_rdo(wb, periodo, rdo)

    # La huella cubre lo declarado, no el instante de generación: dos
    # exportaciones del mismo periodo con los mismos datos coinciden.
    semilla = "|".join(
        f"{f['id_frasco']}:{f['saldo_inicial_g']}:{f['saldo_final_g']}:{f['cantidad_g']}"
        for f in rdo
    )
    huella = hashlib.sha256(semilla.encode("utf-8")).hexdigest()
    _hoja_control(wb, periodo, usuario, huella, len(rdo))

    _hoja_tabla(
        wb,
        "MOVIMIENTOS",
        ["Folio", "Fecha operación", "Tipo", "Motivo", "Código insumo",
         "Insumo", "Código SUNAT", "Presentación", "Estado físico", "Lote",
         "Frasco", "Unidad origen", "Cantidad origen", "Densidad g/mL",
         "Cantidad (g)", "Saldo resultante (g)", "Curso", "Custodio",
         "Registrado por"],
        [[m["folio"], m["fecha_operacion"], m["tipo_movimiento"], m["motivo"],
          m["insumo_codigo"], m["nombre_comercial"], m["codigo_bf_sunat"] or "",
          m["id_presentacion"], m["estado_fisico"], m["numero_lote"] or "",
          m["id_frasco"], m["unidad_registrada"] or "",
          float(m["cantidad_registrada"]) if m["cantidad_registrada"] is not None else "",
          float(m["densidad_aplicada"]) if m["densidad_aplicada"] is not None else "",
          float(m["cantidad_g"]), float(m["saldo_resultante_g"]),
          m["curso"] or "", m["responsable"], m["registrado_por"]]
         for m in movimientos],
    )

    _hoja_tabla(
        wb,
        "RECONCILIACION",
        ["Código SUNAT", "Insumo", "Frascos", "Saldo inicial (kg)",
         "Entradas (kg)", "Salidas (kg)", "Ajustes (kg)", "Saldo teórico (kg)",
         "Saldo físico (kg)", "Diferencia (kg)", "Estado"],
        [[r["codigo_bf_sunat"] or "", r["nombre_comercial"], r["frascos"],
          _kg(r["inicial_g"]), _kg(r["entradas_g"]), _kg(r["salidas_g"]),
          _kg(r["ajustes_g"]), _kg(r["teorico_g"]), _kg(r["fisico_g"]),
          _kg(r["diferencia_g"]), r["estado"]]
         for r in reconciliacion],
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Endpoint ───────────────────────────────────────────────────────────────


def _mes_actual() -> str:
    return dt.date.today().strftime("%Y-%m")


@router.get(
    "/rm04",
    summary="Registro Mensual SUNAT (RM04)",
    description=(
        "Exporta el libro del periodo: la hoja RDO con el formato oficial que "
        "se presenta a SUNAT, más CONTROL, MOVIMIENTOS y RECONCILIACION como "
        "sustento de auditoría."
    ),
)
def exportar_rm04(
    periodo: Annotated[
        str,
        Query(description="Periodo en formato AAAA-MM", pattern=r"^\d{4}-(0[1-9]|1[0-2])$"),
    ] = None,  # type: ignore[assignment]
    connection: Connection = Depends(get_connection),
    current_user: CurrentUser = Depends(get_current_user),
) -> Response:
    periodo = periodo or _mes_actual()
    parametros = {"periodo": periodo}

    columnas_rdo = (
        "nombre_comercial", "grado_pureza", "codigo_bf_sunat", "id_frasco",
        "tipo_envase", "capacidad", "unidad", "equivalencia_g",
        "saldo_inicial_g", "saldo_final_g", "fecha_operacion",
        "tipo_movimiento", "curso", "cantidad_g",
    )
    rdo = [
        dict(zip(columnas_rdo, fila))
        for fila in connection.execute(SQL_RDO, parametros).fetchall()
    ]

    columnas_mov = (
        "folio", "fecha_operacion", "tipo_movimiento", "motivo", "insumo_codigo",
        "nombre_comercial", "codigo_bf_sunat", "id_presentacion", "estado_fisico",
        "numero_lote", "id_frasco", "unidad_registrada", "cantidad_registrada",
        "densidad_aplicada", "cantidad_g", "saldo_resultante_g", "curso",
        "responsable", "registrado_por",
    )
    movimientos = [
        dict(zip(columnas_mov, fila))
        for fila in connection.execute(SQL_MOVIMIENTOS, parametros).fetchall()
    ]

    reconciliacion = []
    for fila in connection.execute(SQL_RECONCILIACION, parametros).fetchall():
        (codigo, nombre, frascos, inicial, entradas, salidas, ajustes, fisico) = fila
        teorico = (Decimal(str(inicial)) + Decimal(str(entradas))
                   - Decimal(str(salidas)) + Decimal(str(ajustes)))
        diferencia = Decimal(str(fisico)) - teorico
        reconciliacion.append({
            "codigo_bf_sunat": codigo,
            "nombre_comercial": nombre,
            "frascos": frascos,
            "inicial_g": inicial,
            "entradas_g": entradas,
            "salidas_g": salidas,
            "ajustes_g": ajustes,
            "teorico_g": teorico,
            "fisico_g": fisico,
            "diferencia_g": diferencia,
            # Cuatro decimales en gramos es la escala del kardex: por debajo
            # de eso la diferencia es ruido de redondeo, no un descuadre.
            "estado": "CONFORME" if abs(diferencia) < Decimal("0.0001") else "PENDIENTE",
        })

    contenido = generar_excel_rm04(
        periodo=periodo,
        rdo=rdo,
        movimientos=movimientos,
        reconciliacion=reconciliacion,
        usuario=current_user.codigo_institucional or current_user.nombre,
    )

    nombre_archivo = f"RM04_SUNAT_{re.sub(r'[^0-9-]', '', periodo)}.xlsx"
    return Response(
        content=contenido,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )
