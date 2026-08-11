"""El libro RM04 se declara ante SUNAT: su forma es parte del contrato.

Estas pruebas fijan la estructura de la hoja `RDO` contra la plantilla
oficial —cabecera combinada de cinco filas, orden y título de cada
columna— para que un cambio de formato tenga que ser deliberado y no se
cuele en una refactorización.
"""

import io

import openpyxl
import pytest

from app.modules.inventario.reportes_rm04 import (
    _grado,
    _presentaciones,
    _unidad_comercial,
    generar_excel_rm04,
)


def _fila_rdo(**cambios):
    base = {
        "nombre_comercial": "Ácido Clorhídrico",
        "grado_pureza": "EMSURE, ACS, ISO, Reag. Ph Eur",
        "codigo_bf_sunat": "000111",
        "id_frasco": "IQF0102-111-98",
        "tipo_envase": "Botella de vidrio",
        "capacidad": 2.5,
        "unidad": "L",
        "equivalencia_g": 2950,
        "saldo_inicial_g": 1794.5,
        "saldo_final_g": 1752.81,
        "fecha_operacion": None,
        "tipo_movimiento": None,
        "curso": None,
        "cantidad_g": None,
    }
    base.update(cambios)
    return base


def _libro(rdo=None, movimientos=None, reconciliacion=None, periodo="2025-09"):
    contenido = generar_excel_rm04(
        periodo=periodo,
        rdo=rdo if rdo is not None else [_fila_rdo()],
        movimientos=movimientos or [],
        reconciliacion=reconciliacion or [],
        usuario="PRUEBA",
    )
    assert isinstance(contenido, bytes) and contenido
    return openpyxl.load_workbook(io.BytesIO(contenido))


def test_el_libro_trae_la_hoja_oficial_primero_y_el_sustento_detras():
    wb = _libro()
    assert wb.sheetnames == ["RDO", "CONTROL", "MOVIMIENTOS", "RECONCILIACION"]


def test_la_cabecera_rdo_reproduce_la_plantilla_sunat():
    """Cada título y cada combinación sale del archivo oficial de SUNAT."""
    ws = _libro()["RDO"]

    esperados = {
        "A1": "N°",
        "B1": "Nombre Comercial del Producto",
        "C1": "Grado del Producto",
        "D1": "Alta SUNAT",
        "E1": "Código Interno Laboratorio",
        "F1": "Unidad Comercial",
        "H1": "Unidad de medida de control",
        "J1": "Stock existente al",
        "J3": "Nº Botellas/ sacos",
        "K3": "Cant. KG",
        "L1": "Movimiento del reactivo",
        "L3": "Fecha",
        "M3": "Tipo de Registro",
        "N3": "Curso",
        "O3": "Consumo",
        "O4": "Cantidad Neta (Kg)",
        "Q1": "Stock existente al",
        "Q3": "CANT. Kg",
        "R3": "Nº Botellas/ Sacos",
    }
    for celda, titulo in esperados.items():
        assert ws[celda].value == titulo, celda

    combinadas = {str(rango) for rango in ws.merged_cells.ranges}
    for rango in ("A1:A5", "J1:K1", "J3:J5", "L1:P1", "O3:P3", "Q1:R1", "S1:S5"):
        assert rango in combinadas, rango

    # Los datos empiezan en la fila 6, debajo de la cabecera de cinco filas.
    assert ws.freeze_panes == "A6"


def test_la_cabecera_fecha_el_periodo_declarado():
    ws = _libro(periodo="2025-09")["RDO"]
    assert ws["J2"].value == "01 DE SETIEMBRE DEL 2025"
    assert ws["Q2"].value == "30 DE SETIEMBRE DEL 2025"
    assert ws["L2"].value == "SETIEMBRE"


def test_febrero_bisiesto_cierra_el_29():
    ws = _libro(periodo="2024-02")["RDO"]
    assert ws["Q2"].value == "29 DE FEBRERO DEL 2024"


def test_un_frasco_sin_movimiento_declara_existencia_igual():
    ws = _libro()["RDO"]
    assert ws["A6"].value == 1
    assert ws["E6"].value == "IQF0102-111-98"
    assert ws["D6"].value == "000111"
    assert ws["K6"].value == pytest.approx(1.7945)   # stock inicial en kg
    assert ws["Q6"].value == pytest.approx(1.75281)  # stock final en kg
    # EN-004: la celda sin dato queda vacía, no se rellena con cero.
    assert ws["L6"].value in (None, "")               # sin fecha de movimiento
    assert ws["O6"].value in (None, "")               # sin consumo


def test_el_codigo_sunat_conserva_los_ceros_a_la_izquierda():
    ws = _libro()["RDO"]
    assert ws["D6"].value == "000111"
    assert ws.cell(6, 4).number_format == "@"


def test_solo_la_salida_se_declara_como_consumo():
    filas = [
        _fila_rdo(tipo_movimiento="SALIDA", cantidad_g=41.69,
                  fecha_operacion="2025-09-26", curso="INVESTIGACIÓN"),
        _fila_rdo(id_frasco="IQF0102-115-99", tipo_movimiento="ENTRADA",
                  cantidad_g=500, fecha_operacion="2025-09-10"),
    ]
    ws = _libro(rdo=filas)["RDO"]

    assert ws["M6"].value == "4"                       # SALIDA → uso
    assert ws["O6"].value == pytest.approx(0.04169)    # consumo en kg
    assert ws["N6"].value == "INVESTIGACIÓN"

    assert ws["M7"].value == "1"                       # ENTRADA → ingreso
    assert ws["O7"].value in (None, "")                # un ingreso no es consumo
    assert ws["P7"].value == 0


def test_un_tipo_fuera_del_catalogo_no_se_disfraza_de_ingreso_ni_de_uso():
    """Un AJUSTE mal codificado como 1 o 4 sería una declaración falsa."""
    ws = _libro(rdo=[_fila_rdo(tipo_movimiento="AJUSTE", cantidad_g=10)])["RDO"]
    assert ws["M6"].value == "AJUSTE"


def test_dos_movimientos_del_mismo_frasco_comparten_numero():
    filas = [
        _fila_rdo(tipo_movimiento="SALIDA", cantidad_g=10, fecha_operacion="2025-09-05"),
        _fila_rdo(tipo_movimiento="SALIDA", cantidad_g=20, fecha_operacion="2025-09-20"),
        _fila_rdo(id_frasco="IQF0102-115-99"),
    ]
    ws = _libro(rdo=filas)["RDO"]
    assert [ws.cell(fila, 1).value for fila in (6, 7, 8)] == [1, 1, 2]


def test_la_huella_depende_de_lo_declarado_y_no_del_instante():
    """Dos exportaciones del mismo periodo y mismos datos deben coincidir."""
    filas = [_fila_rdo()]
    primera = _libro(rdo=filas)["CONTROL"]["B9"].value
    segunda = _libro(rdo=filas)["CONTROL"]["B9"].value
    assert primera == segunda

    distinta = _libro(rdo=[_fila_rdo(saldo_final_g=999)])["CONTROL"]["B9"].value
    assert distinta != primera


def test_control_declara_periodo_establecimiento_y_filas():
    ws = _libro()["CONTROL"]
    valores = {ws.cell(fila, 1).value: ws.cell(fila, 2).value
               for fila in range(2, ws.max_row + 1)}
    assert valores["periodo"] == "2025-09"
    assert valores["establecimiento_codigo"] == "ULIMA-DOCIMASIA"
    assert valores["tipo_reporte"] == "RM04_SUNAT"
    assert valores["filas_declaradas"] == "1"


def test_reconciliacion_marca_pendiente_la_diferencia_real():
    reconciliacion = [
        {"codigo_bf_sunat": "000111", "nombre_comercial": "Ácido Clorhídrico",
         "frascos": 2, "inicial_g": 1000, "entradas_g": 0, "salidas_g": 100,
         "ajustes_g": 0, "teorico_g": 900, "fisico_g": 900,
         "diferencia_g": 0, "estado": "CONFORME"},
        {"codigo_bf_sunat": "000115", "nombre_comercial": "Ácido Clorhídrico",
         "frascos": 1, "inicial_g": 500, "entradas_g": 0, "salidas_g": 0,
         "ajustes_g": 0, "teorico_g": 500, "fisico_g": 480,
         "diferencia_g": -20, "estado": "PENDIENTE"},
    ]
    ws = _libro(reconciliacion=reconciliacion)["RECONCILIACION"]
    assert ws["A2"].value == "000111"
    assert ws["K2"].value == "CONFORME"
    assert ws["J3"].value == pytest.approx(-0.02)  # la diferencia va en kg
    assert ws["K3"].value == "PENDIENTE"


# ── Normalizaciones hacia el vocabulario de la plantilla ───────────────────

@pytest.mark.parametrize("crudo, esperado", [
    ("EMSURE, ACS, ISO, Reag. Ph Eur", "ACS"),
    ("pro analysi / zur Analyse / GR", "PA"),
    ("Q.P.", "PA"),
    ("Reactivo analitico", "PA"),
    (None, ""),
])
def test_el_grado_se_reduce_al_vocabulario_de_la_plantilla(crudo, esperado):
    assert _grado(crudo) == esperado


@pytest.mark.parametrize("crudo, esperado", [
    ("Botella de vidrio", "BOT"),
    ("botella de PLASTICO", "BOT"),
    ("saco de papel", "SACO"),
    ("envase de plástico", "ENV"),
    (None, ""),
])
def test_la_unidad_comercial_usa_el_codigo_de_la_plantilla(crudo, esperado):
    assert _unidad_comercial(crudo) == esperado


def test_sin_equivalencia_no_se_inventa_una_fraccion_de_envase():
    assert _presentaciones(500, None) == ""
    assert _presentaciones(500, 0) == ""
    assert _presentaciones(1475, 2950) == pytest.approx(0.5)
